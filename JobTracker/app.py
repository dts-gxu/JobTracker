from dotenv import load_dotenv
load_dotenv()

from flask import Flask, render_template, request, redirect, url_for, jsonify, flash, send_file, make_response, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, User, Application
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
import os
import logging
from logging.handlers import RotatingFileHandler
from config import config
import urllib.parse

def create_app(config_name=None):
    """Application factory"""
    app = Flask(__name__)
    
    config_name = config_name or os.environ.get('FLASK_ENV', 'development')
    app.config.from_object(config[config_name])
    
    # Session config
    app.config['SESSION_COOKIE_HTTPONLY'] = False
    app.config['SESSION_COOKIE_SECURE'] = False
    app.config['SESSION_COOKIE_SAMESITE'] = None
    app.config['PERMANENT_SESSION_LIFETIME'] = 3600

    # 初始化扩展
    db.init_app(app)
    
    # 配置日志
    if not app.debug and not app.testing:
        if not os.path.exists('logs'):
            os.mkdir('logs')
        file_handler = RotatingFileHandler('logs/app.log', maxBytes=10240000, backupCount=10)
        file_handler.setFormatter(logging.Formatter(
            '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
        ))
        file_handler.setLevel(logging.INFO)
        app.logger.addHandler(file_handler)
        app.logger.setLevel(logging.INFO)
        app.logger.info('JobTracker started')
    
    return app

app = create_app()

# 初始化Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '请先登录访问该页面'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# 创建数据库表
with app.app_context():
    db.create_all()

@app.route('/register', methods=['GET', 'POST'])
def register():
    """用户注册"""
    if request.method == 'POST':
        try:
            username = request.form.get('username', '').strip()
            email = request.form.get('email', '').strip()
            password = request.form.get('password', '').strip()
            confirm_password = request.form.get('confirm_password', '').strip()
            real_name = request.form.get('real_name', '').strip()
            
            # 验证必填字段
            if not username or not email or not password:
                flash('用户名、邮箱和密码都是必填项', 'error')
                return render_template('register.html')
            
            # 验证密码确认
            if password != confirm_password:
                flash('两次输入的密码不一致', 'error')
                return render_template('register.html')
            
            # 检查用户名是否已存在
            if User.query.filter_by(username=username).first():
                flash('用户名已存在，请选择其他用户名', 'error')
                return render_template('register.html')
            
            # 检查邮箱是否已存在
            if User.query.filter_by(email=email).first():
                flash('邮箱已被注册，请使用其他邮箱', 'error')
                return render_template('register.html')
            
            # 创建新用户
            user = User(
                username=username,
                email=email,
                real_name=real_name,
                phone=request.form.get('phone', '').strip(),
                target_position=request.form.get('target_position', '').strip(),
                graduation_year=request.form.get('graduation_year') or None,
                major=request.form.get('major', '').strip(),
                school=request.form.get('school', '').strip()
            )
            user.set_password(password)
            
            db.session.add(user)
            db.session.commit()
            
            flash('注册成功！请登录', 'success')
            return redirect(url_for('login'))
            
        except Exception as e:
            flash(f'注册失败：{str(e)}', 'error')
            db.session.rollback()
    
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """User login"""
    if current_user.is_authenticated:
        return redirect('/')
        
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        if not username or not password:
            flash('请输入用户名和密码', 'error')
            return render_template('login.html')
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            # 更新最后登录时间
            user.last_login = datetime.utcnow()
            db.session.commit()
            
            login_user(user, remember=request.form.get('remember_me'))
            return redirect('/')
        else:
            flash('用户名或密码错误', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    """User logout"""
    logout_user()
    session.clear()
    response = make_response(redirect(url_for('login')))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    response.set_cookie('session', '', expires=0, path='/')
    response.set_cookie('remember_token', '', expires=0, path='/')
    
    return response

@app.route('/clear-session')
def clear_session():
    """Emergency session clear"""
    session.clear()
    logout_user()
    response = make_response(redirect(url_for('login')))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.set_cookie('session', '', expires=0, path='/')
    response.set_cookie('remember_token', '', expires=0, path='/')
    return response

@app.route('/force-clear')
def force_clear():
    """Force clear all sessions and cache"""
    session.clear()
    logout_user()
    
    response = make_response(redirect(url_for('login')))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
    response.headers['Pragma'] = 'no-cache' 
    response.headers['Expires'] = '0'
    
    response.set_cookie('session', '', expires=0, path='/', domain=None)
    response.set_cookie('remember_token', '', expires=0, path='/', domain=None)
    response.set_cookie('_flashes', '', expires=0, path='/', domain=None)
    
    return response

@app.route('/')
def index():
    """Homepage"""
    if not current_user.is_authenticated:
            return render_template('index.html', 
                             applications=[], 
                             stats={'total': 0, 'by_status': {}},
                             is_logged_in=False,
                             user=None)
    
    try:
        status_filter = request.args.get('status', '')
        search_query = request.args.get('search', '')
        query = Application.query.filter_by(user_id=current_user.id)
        
        if status_filter:
            query = query.filter(Application.status == status_filter)
        
        if search_query:
            query = query.filter(
                (Application.company_name.contains(search_query)) |
                (Application.position_name.contains(search_query))
            )
        
        applications = query.order_by(Application.created_at.desc()).all()
        stats = current_user.get_stats()
        
        return render_template('index.html', 
                             applications=applications, 
                             status_options=getattr(Application, 'STATUS_OPTIONS', ['已投递', '面试中', '已拒绝', 'Offer']),
                             current_status=status_filter,
                             search_query=search_query,
                             stats=stats,
                             is_logged_in=True,
                             user=current_user)
    except Exception as e:
        app.logger.error(f"首页加载错误: {str(e)}")
        flash('数据加载失败，请重试', 'error')
        return render_template('index.html', 
                             applications=[], 
                             stats={'total': 0, 'by_status': {}},
                             is_logged_in=True,
                             user=current_user)

@app.route('/add', methods=['GET', 'POST'])
@login_required
def add_application():
    """添加新的投递记录"""
    if request.method == 'POST':
        try:
            # 确保用户已登录
            if not current_user.is_authenticated:
                flash('请先登录', 'error')
                return redirect(url_for('login'))
            
            # 获取表单数据
            company = request.form.get('company_name', '').strip()
            position = request.form.get('position_name', '').strip()
            apply_date_str = request.form.get('apply_date', '').strip()
            status = request.form.get('status', '已投递')
            notes = request.form.get('notes', '').strip()
            
            # 验证必填字段
            if not company or not position:
                flash('公司名称和职位名称都是必填项', 'error')
                return render_template('add.html')
            
            # 处理日期
            if apply_date_str:
                try:
                    apply_date = datetime.strptime(apply_date_str, '%Y-%m-%d').date()
                except ValueError:
                    apply_date = date.today()
            else:
                apply_date = date.today()
            
            # 创建新记录
            application = Application(
                user_id=current_user.id,
                company_name=company,
                position_name=position,
                apply_date=apply_date,
                status=status,
                notes=notes
            )
            
            db.session.add(application)
            db.session.commit()
            
            flash('投递记录添加成功！', 'success')
            return redirect(url_for('index'))
            
        except Exception as e:
            flash(f'添加失败：{str(e)}', 'error')
            db.session.rollback()
            app.logger.error(f"添加投递记录失败: {str(e)}")
    
    return render_template('add.html', 
                         today=date.today().strftime('%Y-%m-%d'),
                         status_options=Application.STATUS_OPTIONS)

@app.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_application(id):
    """编辑投递记录"""
    application = Application.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    
    if request.method == 'POST':
        try:
            # 更新数据
            application.company_name = request.form.get('company_name', '').strip()
            application.position_name = request.form.get('position_name', '').strip()
            apply_date_str = request.form.get('apply_date', '').strip()
            application.status = request.form.get('status', application.status)
            application.notes = request.form.get('notes', '').strip()
            
            # 验证必填字段
            if not application.company_name or not application.position_name or not apply_date_str:
                flash('公司名称、职位名称和投递日期都是必填项', 'error')
                return render_template('edit.html', 
                                     application=application, 
                                     status_options=Application.STATUS_OPTIONS,
                                     channel_options=Application.CHANNEL_OPTIONS)
            
            # 转换日期格式
            application.apply_date = datetime.strptime(apply_date_str, '%Y-%m-%d').date()
            application.updated_at = datetime.utcnow()
            
            # 更新其他字段
            application.salary_min = request.form.get('salary_min') or None
            application.salary_max = request.form.get('salary_max') or None
            application.work_location = request.form.get('work_location', '').strip()
            application.apply_channel = request.form.get('apply_channel', '').strip()
            application.referrer = request.form.get('referrer', '').strip()
            
            # 处理面试时间
            interview_time_str = request.form.get('interview_time', '').strip()
            if interview_time_str:
                application.interview_time = datetime.strptime(interview_time_str, '%Y-%m-%dT%H:%M')
            else:
                application.interview_time = None
            
            db.session.commit()
            flash('投递记录更新成功！', 'success')
            return redirect(url_for('index'))
            
        except ValueError:
            flash('日期时间格式错误，请重新选择', 'error')
        except Exception as e:
            flash(f'更新失败：{str(e)}', 'error')
            db.session.rollback()
    
    return render_template('edit.html', 
                         application=application, 
                         status_options=Application.STATUS_OPTIONS,
                         channel_options=Application.CHANNEL_OPTIONS)

@app.route('/delete/<int:id>', methods=['GET', 'POST'])
@login_required
def delete_application(id):
    """删除投递记录"""
    # 如果是GET请求，检查确认参数
    if request.method == 'GET':
        if not request.args.get('confirm'):
            flash('无效的删除请求', 'error')
            return redirect(url_for('index'))
    
    try:
        application = Application.query.filter_by(id=id, user_id=current_user.id).first_or_404()
        company_name = application.company_name
        position_name = application.position_name
        
        db.session.delete(application)
        db.session.commit()
        flash(f'投递记录删除成功！（{company_name} - {position_name}）', 'success')
    except Exception as e:
        flash(f'删除失败：{str(e)}', 'error')
        db.session.rollback()
    
    return redirect(url_for('index'))

@app.route('/export/excel')
@login_required
def export_excel():
    """导出当前用户的投递记录到Excel"""
    try:
        # 获取当前用户的所有投递记录
        applications = Application.query.filter_by(user_id=current_user.id)\
                                      .order_by(Application.created_at.desc()).all()
        
        # 检查是否有数据
        if not applications:
            flash('暂无投递记录可导出', 'warning')
            return redirect(url_for('index'))
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "投递记录"
        
        # 设置表头
        headers = Application.get_excel_headers()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 写入数据
        for row, app in enumerate(applications, 2):
            data_row = app.to_excel_row()
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col, value=value)
        
        # 调整列宽 - 修复列宽设置
        column_widths = [15, 20, 12, 10, 12, 12, 10, 10, 16, 30, 12]
        for col, width in enumerate(column_widths, 1):
            # 修复：使用正确的列字母计算
            if col <= 26:  # A-Z
                col_letter = chr(64 + col)
            else:  # AA, AB, etc.
                col_letter = chr(64 + col // 26) + chr(64 + col % 26)
            ws.column_dimensions[col_letter].width = width
        
        # 添加统计信息
        stats_row = len(applications) + 3
        ws.cell(row=stats_row, column=1, value="统计信息").font = Font(bold=True)
        
        stats = current_user.get_stats()
        ws.cell(row=stats_row + 1, column=1, value=f"总投递数：{stats['total']}")
        ws.cell(row=stats_row + 2, column=1, value=f"Offer数：{stats['by_status'].get('Offer', 0)}")
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名（简化中文处理）
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_cn = f"{current_user.username}_投递记录_{timestamp}.xlsx"
        
        # 使用send_file方式返回文件
        return send_file(
            output,
            as_attachment=True,
            download_name=filename_cn,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        app.logger.error(f"Excel export error: {str(e)}")
        flash('Excel导出失败，请重试', 'error')
        return redirect(url_for('index'))

@app.route('/profile')
@login_required
def profile():
    """用户个人资料页面"""
    return render_template('profile.html', user=current_user)

@app.route('/api/stats')
@login_required
def api_stats():
    """API接口 - 获取当前用户的统计数据"""
    stats = current_user.get_stats()
    return jsonify(stats)

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('500.html'), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000) 
